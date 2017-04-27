from openpyxl import load_workbook, Workbook
import requests
from bs4 import BeautifulSoup
import re

# Load basic data
wb = load_workbook("taxa.xlsx")
data = wb.get_sheet_by_name("Data")

# Store ID's from table to a list
taxon_id = []
for i in range(6642, 8150):
    taxon_id.append(data.cell(column=1, row=i).value)

# Create new file for the results
results = Workbook()
sheet = results.active

# Get data and store it in new file
for i in range(0, len(taxon_id)):
    taxon_page = requests.get('https://artfakta.artdatabanken.se/taxon/' + str(taxon_id[i]))
    page_content = taxon_page.content
    soup = BeautifulSoup(page_content, "html.parser")

    header = soup.find_all("h2")
    name = header[0].find_all("em")[0].text  # Name of taxon; goes to column "A"
    try:
        taxonomy = soup.find_all("section", {"id":"TaxonomicSection"})
        spans = taxonomy[0].find_all("span")
        synonyms = spans[-1].text  # Synonyms; goes to column "B"

        group = soup.find_all("div", {"id": "tab3_collapseSeven"})  # div which contains most of info

        pattern = re.compile(r"Ekologisk")  # Regex for pulling only food preference data
        food_preference = group[0].find("span", text=pattern).parent.text.strip().replace(" ", "")  # Taxon ecology: goes to column "D"

        ecology = group[0].find_all("span", {"data-toggle": "popover"})
        tree = []  # Taxon habitat; goes to column "C"
        for item in ecology:
            tree.append(item.text)
        tree = [s.strip() for s in tree]

        column_a = sheet.cell(column=1, row=i+2, value=name)
        column_b = sheet.cell(column=2, row=i+2, value=synonyms)
        column_c = sheet.cell(column=3, row=i+2, value=str(tree))
        column_d = sheet.cell(column=4, row=i+2, value=food_preference)

        results.save("rest_of_data_2.xlsx")

        print("Tehtud rida " + str(i))  # Show progress
    except:
        column_a = sheet.cell(column=1, row=i+2, value=name)
        column_b = sheet.cell(column=2, row=i+2, value="-")
        column_c = sheet.cell(column=3, row=i+2, value="-")
        column_d = sheet.cell(column=4, row=i+2, value="-")

        results.save("rest_of_data_2.xlsx")

        print("Tehtud rida " + str(i))
        pass
